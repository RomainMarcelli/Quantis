#!/usr/bin/env python3
"""Build kpi_review.xlsx from /tmp/kpi-data.json.

Two sheets:
  1. "Registre KPIs" — one row per KPI
  2. "Scénarios Simulation" — one row per scenario

Cells with subjective content (tooltips, benchmarks, suggested questions,
thresholds, lever bounds) are highlighted yellow as a review checklist.
Objective cells (id, label, formula derived from code, phase, source layer)
stay white.

Usage:
    python3 scripts/build-kpi-review-xlsx.py [--output PATH]
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
HEADER = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


# Headers per sheet — order matters for column writing.
KPI_HEADERS: list[tuple[str, bool]] = [
    # (column label, needs_review)
    ("ID", False),
    ("Label", False),
    ("Catégorie", False),
    ("Formule (français)", False),
    ("Formule (code)", False),
    ("Unité", False),
    ("Tooltip Explication", True),
    ("Tooltip Bon Signe", True),
    ("Tooltip Mauvais Signe", True),
    ("Benchmark", True),
    ("Question si bon", True),
    ("Question si mauvais", True),
    ("Seuil danger", True),
    ("Seuil warning", True),
    ("Seuil good", True),
    ("Leviers de simulation", True),
    ("Phase (CT/MT/LT)", False),
    ("Source (accounting/banking/both)", False),
]

SCENARIO_HEADERS: list[tuple[str, bool]] = [
    ("ID", False),
    ("Label", False),
    ("Description", True),
    ("Leviers", True),
    ("KPIs impactés", False),
]


def kpi_row(kpi: dict) -> list[object]:
    tooltip = kpi.get("tooltip") or {}
    sq = kpi.get("suggestedQuestions") or {}
    thresholds = kpi.get("thresholds") or {}
    simulation = kpi.get("simulation") or {}
    levers = ", ".join(simulation.get("levers", []))
    return [
        kpi.get("id", ""),
        kpi.get("label", ""),
        kpi.get("category", ""),
        kpi.get("formula", ""),
        kpi.get("formulaCode", ""),
        kpi.get("unit", ""),
        tooltip.get("explanation", ""),
        tooltip.get("goodSign", ""),
        tooltip.get("badSign", ""),
        tooltip.get("benchmark", ""),
        sq.get("whenGood", ""),
        sq.get("whenBad", ""),
        thresholds.get("danger", ""),
        thresholds.get("warning", ""),
        thresholds.get("good", ""),
        levers,
        kpi.get("phase", ""),
        kpi.get("sourceLayer", ""),
    ]


def scenario_row(scenario: dict) -> list[object]:
    levers_repr_lines = []
    for lever in scenario.get("levers", []):
        hidden_tag = " (hidden)" if lever.get("hidden") else ""
        levers_repr_lines.append(
            f"• {lever.get('variableCode', '')} — {lever.get('label', '')} "
            f"[{lever.get('type', '')}; "
            f"min={lever.get('min')}, max={lever.get('max')}, "
            f"step={lever.get('step')}, default={lever.get('defaultDelta')}]"
            f"{hidden_tag}"
        )
    affected = ", ".join(scenario.get("affectedKpis", []))
    return [
        scenario.get("id", ""),
        scenario.get("label", ""),
        scenario.get("description", ""),
        "\n".join(levers_repr_lines),
        affected,
    ]


def write_sheet(ws, headers: list[tuple[str, bool]], rows: list[list[object]]) -> None:
    # Header row
    for col_idx, (header, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows + per-cell yellow flag
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, ((_, needs_review), value) in enumerate(zip(headers, row_values), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP
            if needs_review and value not in (None, ""):
                cell.fill = YELLOW

    # Column widths : header length × 1.4 ; clamp 12-60
    for col_idx, (header, _) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Compute max content width across rows for this column
        max_len = len(header)
        for r in rows:
            v = r[col_idx - 1]
            if v is None:
                continue
            text = str(v)
            for line in text.split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        width = min(60, max(12, int(max_len * 1.05) + 2))
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28
    # Row heights : auto-fit by counting newlines + length
    for row_idx, row_values in enumerate(rows, start=2):
        max_lines = 1
        for v in row_values:
            if v in (None, ""):
                continue
            text = str(v)
            # Estimate wrap : assume 60 chars per line within wrapped column
            for line in text.split("\n"):
                wraps = max(1, (len(line) // 55) + 1)
                if wraps > max_lines:
                    max_lines = wraps
        if max_lines > 1:
            ws.row_dimensions[row_idx].height = min(180, 18 * max_lines + 4)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="/tmp/kpi-data.json")
    parser.add_argument("--output", default="/tmp/kpi_review.xlsx")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))

    wb = Workbook()
    # Sheet 1
    ws1 = wb.active
    ws1.title = "Registre KPIs"
    kpi_rows = [kpi_row(k) for k in data["kpis"]]
    write_sheet(ws1, KPI_HEADERS, kpi_rows)
    # Freeze header
    ws1.freeze_panes = "B2"

    # Sheet 2
    ws2 = wb.create_sheet("Scénarios Simulation")
    scenario_rows = [scenario_row(s) for s in data["scenarios"]]
    write_sheet(ws2, SCENARIO_HEADERS, scenario_rows)
    ws2.freeze_panes = "B2"

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(
        f"OK : {out}\n"
        f"  - {len(kpi_rows)} KPI(s)\n"
        f"  - {len(scenario_rows)} scénario(s)\n"
        f"  - colonnes en jaune = à reviewer"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
